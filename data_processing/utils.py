import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

def clean_price_column(df: pd.DataFrame, column_name: str) -> pd.DataFrame:
    df[column_name] = df[column_name].astype(str)  # Convert to string
    df[column_name] = df[column_name].str.replace('.', '', regex=False)  # Remove periods
    df[column_name] = df[column_name].astype(int)  # Convert to integer

    return df

# Function to save processed_data to Excel
def save_processed_data_to_excel(writer, processed_data):
    for sheet_name, data in processed_data.items():
        data.to_excel(writer, sheet_name=sheet_name, index=False)

def apply_styles(sheets):
    for sheet_name, ws in sheets.items():
        # Apply styling to the header row
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        # Set column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
