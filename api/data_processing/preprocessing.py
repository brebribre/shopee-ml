import pandas as pd
from utils import clean_price_column
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

def process_single_sheet(sheet_name, xls, columns_to_read):
    # Load data from the specified sheet
    data = pd.read_excel(xls, sheet_name=sheet_name, usecols=columns_to_read, dtype=str)

    # Only process 'finished' orders
    data = data[data['Status Pesanan'] == 'Selesai']

    # Specify missing data
    data['Nomor Referensi SKU'] = data['Nomor Referensi SKU'].fillna('UNKNOWN')

    # Remove commas and convert the columns to float from IDR values
    data = clean_price_column(data, 'Harga Awal')
    data = clean_price_column(data, 'Harga Setelah Diskon')
    data = clean_price_column(data, 'Total Harga Produk')
    data['Jumlah'] = data['Jumlah'].astype(int)

    # Group the data by 'Nomor Referensi SKU' and 'Nama Produk'
    grouped_data = data.groupby(['Nomor Referensi SKU', 'Nama Produk'], as_index=False).agg({
        'Jumlah': 'sum',
        'Total Harga Produk': 'sum'
    })

    # Add a total row at the end of the grouped data
    total_row = {
        'Nomor Referensi SKU': 'TOTAL',
        'Nama Produk': '',
        'Jumlah': grouped_data['Jumlah'].sum(),
        'Total Harga Produk': grouped_data['Total Harga Produk'].sum()
    }

    # Append the total row to the grouped data
    grouped_data = pd.concat([grouped_data, pd.DataFrame(total_row, index=[0])])

    # Store the processed data for the sheet
    processed_data[sheet_name] = grouped_data

# Function to save processed_data to Excel
def save_processed_data_to_excel(output_filepath):
    with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
        for sheet_name, data in processed_data.items():
            data.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"Processed data saved to {output_filepath}")

     
def apply_styles(output_file):
    # Load the workbook to apply formatting
    wb = load_workbook(output_file)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Apply custom styles for the header row
        for cell in ws[1]:
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow background
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Adjust column widths
        ws.column_dimensions['A'].width = 20  # Nomor Referensi SKU
        ws.column_dimensions['B'].width = 80  # Nama Produk
        ws.column_dimensions['C'].width = 10  # Jumlah
        ws.column_dimensions['D'].width = 20  # Total Harga Produk

    # Save the workbook with the applied styles
    wb.save(output_file)
    print(f"Styles applied and file saved as '{output_file}'.")


input_file = './assets/sep2022-ag2024.xlsx'
output_file = './assets/processed_output.xlsx'

# Load the Excel file
xls = pd.ExcelFile(input_file, engine='openpyxl')

# Create a dictionary to store the processed data for each sheet
processed_data = {}

# Specify the columns you want to extract
columns_to_read = [
     'No. Pesanan', 
     'Status Pesanan', 
     'Nomor Referensi SKU', 
     'Harga Awal', 
     'Harga Setelah Diskon', 
     'Total Harga Produk', 
     'Jumlah', 
     'Nomor Referensi SKU', 
     'Nama Produk'
     ]

# Process a specific sheet (e.g., 'SEPT 2022')
process_single_sheet('AGUSTUS 2024', xls, columns_to_read)

# Save the processed data to an Excel file
save_processed_data_to_excel(output_file)
apply_styles(output_file)