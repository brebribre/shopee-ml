import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

def clean_price_column(df: pd.DataFrame, column_name: str) -> pd.DataFrame:
    """
    This function cleans a price column in a pandas DataFrame by:
    1. Treating the column as a string to preserve periods.
    2. Removing periods (.) from the string.
    3. Converting the cleaned string to integers.

    :param df: The pandas DataFrame containing the data.
    :param column_name: The name of the column to clean.
    :return: The DataFrame with the cleaned column.
    """
    df[column_name] = df[column_name].astype(str)  # Convert to string
    df[column_name] = df[column_name].str.replace('.', '', regex=False)  # Remove periods
    df[column_name] = df[column_name].astype(int)  # Convert to integer

    return df

# Function to save processed_data to Excel
def save_processed_data_to_excel(output_filepath, processed_data):
    with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
        for sheet_name, data in processed_data.items():
            data.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"Processed data saved to {output_filepath}")


def apply_styles(output_file):
    # Load the workbook to apply formatting
    wb = load_workbook(output_file)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Apply custom styles for the header row
        for cell in ws[1]:
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow background
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Adjust column widths
        ws.column_dimensions['A'].width = 20  # Nomor Referensi SKU
        ws.column_dimensions['B'].width = 80  # Nama Produk
        ws.column_dimensions['C'].width = 10  # Jumlah
        ws.column_dimensions['D'].width = 20  # Total Harga Produk

    # Save the workbook with the applied styles
    wb.save(output_file)
    print(f"Styles applied and file saved as '{output_file}'.")

